import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import requests
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from pathlib import Path

# -----------------------------
# SETTINGS
# -----------------------------
EXCEL_FILE = "DTR_Records.xlsx"
ADMIN_PASSWORD = "admin123"
DOWNLOADS_DIR = str(Path.home() / "Downloads")

# -----------------------------
# CREATE EXCEL FILE IF NOT EXISTS
# -----------------------------
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    wb.save(EXCEL_FILE)

# -----------------------------
# INTERNET CHECK
# -----------------------------
def internet_available():
    try:
        requests.get("https://google.com", timeout=3)
        return True
    except:
        return False

# -----------------------------
# GOOGLE SHEETS SYNC (FAST)
# -----------------------------
def sync_to_google():
    try:
        scope = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = ServiceAccountCredentials.from_json_keyfile_name("google_credentials.json", scope)
        client = gspread.authorize(creds)
        sheet_file = client.open("Office DTR Database")

        wb = load_workbook(EXCEL_FILE)
        for emp_id in wb.sheetnames:
            ws = wb[emp_id]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            try:
                emp_sheet = sheet_file.worksheet(emp_id)
            except gspread.WorksheetNotFound:
                emp_sheet = sheet_file.add_worksheet(title=emp_id, rows="1000", cols="10")

            emp_sheet.clear()
            # Batch update all rows at once
            emp_sheet.update(rows)
        status_label.config(text="☁ Synced to Google Sheets ✅", fg="green")
    except Exception as e:
        status_label.config(text=f"⚠ Google Sheets sync failed: {e}", fg="red")
        print("Google Sheets sync error:", e)

# -----------------------------
# CLOCK IN / OUT
# -----------------------------
def log_attendance(clock_type):
    emp_id = emp_id_entry.get().strip()
    if emp_id == "":
        status_label.config(text="❌ Enter Employee ID", fg="red")
        return
    manual_time = time_entry.get().strip()
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = manual_time if manual_time else now.strftime("%I:%M %p")

    wb = load_workbook(EXCEL_FILE)
    if emp_id not in wb.sheetnames:
        ws = wb.create_sheet(emp_id)
        ws.append(["Date", "Time In", "Time Out", "Notes"])
    else:
        ws = wb[emp_id]

    if clock_type == "in":
        ws.append([date_str, time_str, "", "Clocked In"])
        status_label.config(text=f"✅ Time In: {time_str}", fg="green")
    elif clock_type == "out":
        found = False
        for row in range(ws.max_row, 1, -1):
            time_out = ws.cell(row=row, column=3).value
            note = ws.cell(row=row, column=4).value
            if time_out in ("", None) and note == "Clocked In":
                ws.cell(row=row, column=3).value = time_str
                ws.cell(row=row, column=4).value = "Clocked Out"
                status_label.config(text=f"✅ Time Out: {time_str}", fg="blue")
                found = True
                break
        if not found:
            status_label.config(text="❌ No clock-in found", fg="red")

    wb.save(EXCEL_FILE)
    emp_id_entry.delete(0, tk.END)
    time_entry.delete(0, tk.END)

# -----------------------------
# AUTO SYNC
# -----------------------------
def auto_sync():
    if internet_available():
        sync_to_google()
    root.after(60000, auto_sync)  # every 1 minute

# -----------------------------
# GENERATE MONTHLY REPORT
# -----------------------------
def generate_monthly_report(emp_id, year, month):
    try:
        wb = load_workbook(EXCEL_FILE)
        if emp_id not in wb.sheetnames:
            return None
        ws = wb[emp_id]
        report_wb = Workbook()
        report_ws = report_wb.active
        report_ws.title = f"{emp_id}_Report"

        # Headers
        headers = [cell.value for cell in ws[1]]
        report_ws.append(headers)

        # Filter rows by month/year
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_cell = row[0]
            if date_cell:
                dt = datetime.strptime(date_cell, "%Y-%m-%d")
                if dt.year == year and dt.month == month:
                    report_ws.append([str(cell) if cell else "" for cell in row])

        # Save silently to Downloads
        report_filename = os.path.join(DOWNLOADS_DIR, f"{emp_id}_Report_{year}_{month}.xlsx")
        report_wb.save(report_filename)
        return report_filename
    except Exception as e:
        print("Generate report error:", e)
        return None

# -----------------------------
# STUDENT VIEW (READ ONLY)
# -----------------------------
def student_view_report():
    emp_id = student_emp_entry.get().strip()
    month_str = student_month_entry.get().strip()
    year_str = student_year_entry.get().strip()
    if not emp_id or not month_str or not year_str:
        return
    try:
        month = int(month_str)
        year = int(year_str)
    except ValueError:
        return
    wb = load_workbook(EXCEL_FILE)
    if emp_id not in wb.sheetnames:
        return
    ws = wb[emp_id]

    preview_win = tk.Toplevel(root)
    preview_win.title(f"{emp_id} Attendance Preview")
    tree = ttk.Treeview(preview_win, columns=("Date", "Time In", "Time Out", "Notes"), show="headings")
    for col in ("Date", "Time In", "Time Out", "Notes"):
        tree.heading(col, text=col)
        tree.column(col, width=120)
    tree.pack(expand=True, fill=tk.BOTH)

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_cell = row[0]
        if date_cell:
            dt = datetime.strptime(date_cell, "%Y-%m-%d")
            if dt.year == year and dt.month == month:
                tree.insert("", tk.END, values=row)

# -----------------------------
# ADMIN PANEL
# -----------------------------
def open_admin_login():
    login_win = tk.Toplevel(root)
    login_win.title("Admin Login")
    login_win.geometry("400x200")

    tk.Label(login_win, text="Enter Admin Password:", font=("Helvetica", 16)).pack(pady=20)
    pwd_entry = tk.Entry(login_win, font=("Helvetica", 16), show="*")
    pwd_entry.pack(pady=10)

    def check_password():
        if pwd_entry.get() == ADMIN_PASSWORD:
            login_win.destroy()
            open_admin_panel()
        else:
            messagebox.showerror("Error", "Incorrect Password")

    tk.Button(login_win, text="Login", font=("Helvetica", 14), command=check_password).pack(pady=10)

def open_admin_panel():
    admin_win = tk.Toplevel(root)
    admin_win.title("Admin Panel")
    admin_win.geometry("1000x600")

    tk.Label(admin_win, text="Admin Panel - View Employee Records", font=("Helvetica", 24, "bold")).pack(pady=10)

    search_frame = tk.Frame(admin_win)
    search_frame.pack(pady=5)
    tk.Label(search_frame, text="Employee ID:", font=("Helvetica", 14)).grid(row=0, column=0)
    search_entry = tk.Entry(search_frame, font=("Helvetica", 14), width=15)
    search_entry.grid(row=0, column=1, padx=5)
    tk.Label(search_frame, text="Month:", font=("Helvetica", 12)).grid(row=1, column=0, pady=5)
    admin_month_entry = tk.Entry(search_frame, font=("Helvetica", 12), width=5)
    admin_month_entry.grid(row=1, column=1, pady=5)
    tk.Label(search_frame, text="Year:", font=("Helvetica", 12)).grid(row=1, column=2, pady=5)
    admin_year_entry = tk.Entry(search_frame, font=("Helvetica", 12), width=7)
    admin_year_entry.grid(row=1, column=3, pady=5)

    def load_records():
        emp_id = search_entry.get().strip()
        for row in tree.get_children():
            tree.delete(row)
        wb = load_workbook(EXCEL_FILE)
        sheetnames = [emp_id] if emp_id in wb.sheetnames else wb.sheetnames
        for sheet in sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                tree.insert("", tk.END, values=(sheet, *row))

    def generate_report_admin_safe():
        emp_id = search_entry.get().strip()
        month_str = admin_month_entry.get().strip()
        year_str = admin_year_entry.get().strip()
        if not emp_id or not month_str or not year_str:
            return
        try:
            month = int(month_str)
            year = int(year_str)
        except ValueError:
            return
        generate_monthly_report(emp_id, year, month)  # silent download

    tk.Button(search_frame, text="Load Records", font=("Helvetica", 12), command=load_records).grid(row=0, column=2, padx=5)
    tk.Button(search_frame, text="Generate Report", font=("Helvetica", 12), command=generate_report_admin_safe).grid(row=0, column=3, padx=5)

    columns = ("Employee ID", "Date", "Time In", "Time Out", "Notes")
    tree = ttk.Treeview(admin_win, columns=columns, show="headings")
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=150)
    tree.pack(expand=True, fill=tk.BOTH, pady=10)

# -----------------------------
# MAIN UI
# -----------------------------
root = tk.Tk()
root.title("Office DTR System")
root.geometry("1280x950")

tk.Label(root, text="📋 OFFICE DTR SYSTEM", font=("Helvetica", 36, "bold")).pack(pady=20)

# Employee clock in/out
tk.Label(root, text="Employee ID", font=("Helvetica", 24)).pack()
emp_id_entry = tk.Entry(root, font=("Helvetica", 28), width=20, justify="center")
emp_id_entry.pack(pady=10)
tk.Label(root, text="Manual Time (Optional)", font=("Helvetica", 18)).pack()
time_entry = tk.Entry(root, font=("Helvetica", 24), width=15, justify="center")
time_entry.pack(pady=10)

button_frame = tk.Frame(root)
button_frame.pack(pady=20)
tk.Button(button_frame, text="CLOCK IN", font=("Helvetica", 28, "bold"), bg="#28a745", fg="white",
          width=12, height=2, command=lambda: log_attendance("in")).grid(row=0, column=0, padx=20)
tk.Button(button_frame, text="CLOCK OUT", font=("Helvetica", 28, "bold"), bg="#dc3545", fg="white",
          width=12, height=2, command=lambda: log_attendance("out")).grid(row=0, column=1, padx=20)

# Student report
tk.Label(root, text="Student Report View (Request from Admin)", font=("Helvetica", 24, "bold")).pack(pady=20)
student_frame = tk.Frame(root)
student_frame.pack(pady=10)
tk.Label(student_frame, text="Employee ID:", font=("Helvetica", 18)).grid(row=0, column=0)
student_emp_entry = tk.Entry(student_frame, font=("Helvetica", 18), width=10)
student_emp_entry.grid(row=0, column=1, padx=10)
tk.Label(student_frame, text="Month (1-12):", font=("Helvetica", 18)).grid(row=0, column=2)
student_month_entry = tk.Entry(student_frame, font=("Helvetica", 18), width=5)
student_month_entry.grid(row=0, column=3, padx=10)
tk.Label(student_frame, text="Year:", font=("Helvetica", 18)).grid(row=0, column=4)
student_year_entry = tk.Entry(student_frame, font=("Helvetica", 18), width=7)
student_year_entry.grid(row=0, column=5, padx=10)
tk.Button(student_frame, text="Request Report", font=("Helvetica", 20, "bold"), bg="#007bff", fg="white",
          command=student_view_report).grid(row=0, column=6, padx=10)

# Admin login
tk.Button(root, text="Admin Panel", font=("Helvetica", 20, "bold"), bg="#6c757d", fg="white",
          command=open_admin_login).pack(pady=20)

status_label = tk.Label(root, text="System Ready", font=("Helvetica", 20))
status_label.pack(pady=20)

# Start auto-sync
auto_sync()
root.mainloop()